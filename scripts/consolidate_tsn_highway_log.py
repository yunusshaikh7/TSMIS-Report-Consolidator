"""Convert TSN district Highway Log PDFs into TSMIS-format Excel and combine.

Reads every district PDF in  input/tsn_highway_log/   (e.g. D01_Highway_Log_TSN.pdf)
Writes per-route workbooks in output/tsn_highway_log/  (tsn_highway_log_d01_route_001.xlsx)
and one combined workbook   output/tsn_highway_log_consolidated.xlsx

The TSN (Transportation System Network) "California State Highway Log"
(report OTM52010) is a fixed-layout PDF listing: a 3-line column-header band
per page, a centered "<district> <county> <route>" group header, one data line
per highway segment (sometimes wrapping onto a second baseline), description
lines *below* the segment they belong to, and "* * Volume Location Totals"
summary lines.

Each per-route output uses the SAME sheet name ("Highway Log") and the SAME 31
columns as the per-route TSMIS Highway Log export, so:
  * the shared XLSX consolidator combines them unchanged (Route prepended from
    the filename), and
  * the combined workbook lines up column-for-column with the consolidated
    TSMIS Highway Log for comparison.
TSN-only data that has no TSMIS column (the ADT traffic figures) is dropped;
TSN description lines are joined into the TSMIS "Description" column.

Parsing is x-position based (the PDF is proportional Helvetica, not
monospaced): every data value is assigned to a column by the horizontal window
its center falls in. The windows are calibrated to the OTM52010 layout and
verified stable across every data row of the sample districts.

Console-free like the other consolidators: progress via events.on_log,
overwrite confirmed through the callback, cancel honored between pages, and a
ConsolidateResult returned. The console UX lives in cli.run_consolidate_cli.
"""
import logging
import re
from pathlib import Path

# pdfplumber wraps pdfminer.six, which can log noisy per-page font warnings;
# parsing is unaffected (see consolidate_ramp_summary).
logging.getLogger("pdfminer").setLevel(logging.ERROR)

try:
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

from consolidate_xlsx_base import consolidate_xlsx
from events import ConsolidateResult, Events
from paths import INPUT_ROOT, OUTPUT_ROOT

INPUT_DIR = INPUT_ROOT / "tsn_highway_log"
CONVERTED_DIR = OUTPUT_ROOT / "tsn_highway_log"   # per-route TSMIS-format workbooks
OUT_PATH = OUTPUT_ROOT / "tsn_highway_log_consolidated.xlsx"

# Must match the TSMIS Highway Log export exactly (sheet name AND header), so
# the converted files consolidate with the same core and the combined workbook
# is column-compatible with the consolidated TSMIS Highway Log.
SHEET_NAME = "Highway Log"
TSMIS_HEADER = [
    "Location", "MI", "N/A", "Cnty Odom", "City", "R/U", "SPD", "TER", "H/G",
    "A/C", "LB T", "LB Lns", "LB F", "LB OT", "LB TR", "LB T-W", "LB IN",
    "LB SH", "Med TCB", "Med Wid", "RB T", "RB Lns", "RB F", "RB IN", "RB SH",
    "RB T-W", "RB OT", "RB SH", "Description", "Date of Rec", "Sig Chg. Date",
]

# Friendly report name for user-facing messages (shown in both the GUI and the
# console, so keep it UI-neutral -- no ".bat" / "menu option" wording).
REPORT_NAME = "TSN Highway Log"

# File pattern the GUI uses to preview how many inputs a folder holds.
INPUT_GLOB = "*.pdf"


# =============================================================================
# PDF layout -- calibrated to the OTM52010 "California State Highway Log"
# =============================================================================

Y_TOLERANCE = 3      # words within this y-distance form one logical line
HEADER_BAND = 56     # everything above this y on a page is page furniture

# (column_key, x_min, x_max): a data word belongs to the column whose window
# contains the word's horizontal CENTER. Order = TSMIS column order; the three
# ADT columns exist in the TSN layout but have no TSMIS counterpart and are
# dropped when rows are written. "Description" has no window -- TSN prints
# descriptions as separate lines below the data row.
COLUMN_WINDOWS = [
    ("location",  0, 50),     # may carry a realignment prefix: "R012.887"
    ("mi",       50, 73),
    ("na",       73, 82),
    ("cnty_odom", 82, 112),
    ("city",    112, 132),
    ("ru",      132, 147),
    ("spd",     147, 160),
    ("ter",     160, 171),
    ("hg",      171, 184),
    ("ac",      184, 197),
    ("lb_t",    197, 208),
    ("lb_lns",  208, 219),
    ("lb_f",    219, 230),
    ("lb_ot",   230, 241),
    ("lb_tr",   241, 253),
    ("lb_tw",   253, 268),
    ("lb_in",   268, 279),
    ("lb_sh",   279, 291),
    ("med_tcb", 291, 308),
    ("med_wid", 308, 326),
    ("rb_t",    326, 338),
    ("rb_lns",  338, 350),
    ("rb_f",    350, 361),
    ("rb_in",   361, 372),
    ("rb_sh",   372, 386),
    ("rb_tw",   386, 398),
    ("rb_ot",   398, 410),
    ("rb_sh2",  410, 424),
    ("adt_back",  424, 448),  # TSN-only (ADT Look Back)   -> dropped
    ("adt_pp",    448, 459),  # TSN-only (ADT P/P flag)    -> dropped
    ("adt_ahead", 459, 486),  # TSN-only (ADT Look Ahead)  -> dropped
    ("rec",     486, 519),
    ("sig",     519, 612),
]

# Row keys in TSMIS column order (Description filled from follow-on lines).
ROW_KEYS = ["location", "mi", "na", "cnty_odom", "city", "ru", "spd", "ter",
            "hg", "ac", "lb_t", "lb_lns", "lb_f", "lb_ot", "lb_tr", "lb_tw",
            "lb_in", "lb_sh", "med_tcb", "med_wid", "rb_t", "rb_lns", "rb_f",
            "rb_in", "rb_sh", "rb_tw", "rb_ot", "rb_sh2",
            "description", "rec", "sig"]

# A segment postmile, optionally with a glued realignment prefix ("R012.887")
# and/or a trailing equation suffix ("026.437E"), as printed in the Location
# column (TSMIS prints the same prefixed form).
LOCATION_RE = re.compile(r"^[A-Z]?\d{3}\.\d{3}[A-Z]?$")
# Centered "<district> <county> <route>" group header, e.g. "01 MEN 001".
GROUP_RE = (re.compile(r"^\d{2}$"), re.compile(r"^[A-Z]{2,4}$"),
            re.compile(r"^\d{1,3}[A-Z]?$"))
DISTRICT_LINE_RE = re.compile(r"^District\s+0?(\d{1,2})$", re.IGNORECASE)
DISTRICT_FROM_NAME = re.compile(r"D(\d{1,2})", re.IGNORECASE)


def _lines(page):
    """Group the page's words into logical lines (left-to-right), tolerating
    the 1pt baseline jitter the report's wrapped data rows have."""
    grouped = []                      # [(top, [word, ...]), ...]
    for w in sorted(page.extract_words(), key=lambda w: (w["top"], w["x0"])):
        if grouped and abs(w["top"] - grouped[-1][0]) <= Y_TOLERANCE:
            grouped[-1][1].append(w)
        else:
            grouped.append((w["top"], [w]))
    return [(top, sorted(ws, key=lambda w: w["x0"])) for top, ws in grouped]


def _parse_data_line(words):
    """Map each word of a data line to its column by horizontal center."""
    row = {}
    for w in words:
        center = (w["x0"] + w["x1"]) / 2
        for key, lo, hi in COLUMN_WINDOWS:
            if lo <= center < hi:
                # Two words in one window means the layout shifted -- join them
                # so nothing is silently lost (shows up visibly in the output).
                row[key] = (row[key] + " " + w["text"]) if key in row else w["text"]
                break
    return row


def _norm_route(token):
    """'1' -> '001' (TSMIS zero-pads); suffixed routes ('101U') kept as-is."""
    return token.zfill(3) if token.isdigit() else token.upper()


def parse_pdf(path, events, pdf_name=""):
    """Parse one TSN district Highway Log PDF.

    Returns (district, routes) where routes is {route: [row_dict, ...]} in
    document order. Raises RuntimeError on cancel (caught by the caller).
    """
    district = None
    routes = {}
    route = None
    last_row = None                   # description lines attach to this

    with pdfplumber.open(path) as pdf:
        n_pages = len(pdf.pages)
        for page_no, page in enumerate(pdf.pages, 1):
            if events.is_cancelled():
                return district, None
            if page_no % 25 == 0:
                events.on_log(f"    …page {page_no}/{n_pages}")
            for top, words in _lines(page):
                if top < HEADER_BAND:
                    continue                          # per-page header band
                texts = [w["text"] for w in words]
                first = words[0]

                # "* * Volume Location Totals ..." summary lines.
                if texts[0].startswith("*"):
                    continue

                # Title page: "District 01" pins the district number.
                m = DISTRICT_LINE_RE.match(" ".join(texts))
                if m and district is None:
                    district = m.group(1).zfill(2)
                    continue

                # Centered group header: "<district> <county> <route>".
                if (len(texts) >= 3 and 250 <= first["x0"] <= 305
                        and GROUP_RE[0].match(texts[0])
                        and GROUP_RE[1].match(texts[1])
                        and GROUP_RE[2].match(texts[2])):
                    district = district or texts[0].zfill(2)
                    route = _norm_route(texts[2])
                    routes.setdefault(route, [])
                    last_row = None                   # don't attach across groups
                    continue

                # Data line: starts with a postmile in the Location window.
                if LOCATION_RE.match(texts[0]) and first["x0"] < 50:
                    if route is None:
                        events.on_log(f"    {pdf_name} p{page_no}: data before "
                                      "any route header; line skipped")
                        continue
                    row = _parse_data_line(words)
                    row["description"] = None
                    routes[route].append(row)
                    last_row = row
                    continue

                # Anything else below the band is a description for the
                # previous segment (TSN prints them on their own lines).
                if last_row is not None:
                    text = " ".join(texts)
                    last_row["description"] = (
                        text if not last_row["description"]
                        else last_row["description"] + ", " + text)

    return district, routes


# =============================================================================
# TSMIS-format per-route workbooks
# =============================================================================

def _write_route_workbook(rows, out_path):
    """Write one route's rows as a TSMIS-format Highway Log workbook."""
    header_fill = PatternFill("solid", start_color="305496")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(TSMIS_HEADER)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.freeze_panes = "A2"
    for i, name in enumerate(TSMIS_HEADER, start=1):
        ws.column_dimensions[get_column_letter(i)].width = \
            40 if name == "Description" else 10

    for row in rows:
        ws.append([row.get(k) for k in ROW_KEYS])
    wb.save(out_path)


# =============================================================================
# Entry point
# =============================================================================

def consolidate(events=None, confirm_overwrite=None, input_dir=None, out_path=None):
    """Convert every TSN district Highway Log PDF to TSMIS-format per-route
    workbooks, then combine them all into one workbook (Route column added).

    Console-free: reports progress via events.on_log, asks before overwriting
    through the confirm_overwrite(path)->bool callback, and returns a
    ConsolidateResult. Honors events.is_cancelled() between pages.
    """
    in_dir = Path(input_dir) if input_dir else INPUT_DIR
    out = Path(out_path) if out_path else OUT_PATH
    events = events or Events()
    if not _DEPS_OK:
        return ConsolidateResult(
            status="error",
            message="Required components are missing (pdfplumber, openpyxl).",
        )
    confirm = confirm_overwrite or (lambda _p: True)

    if not in_dir.exists():
        return ConsolidateResult(
            status="error",
            message=(f"The {REPORT_NAME} input folder doesn't exist:\n{in_dir}\n\n"
                     f"Put the district Highway Log PDFs there, then run again."),
        )

    pdfs = sorted(in_dir.glob("*.pdf"))
    if not pdfs:
        return ConsolidateResult(
            status="error",
            message=(f"No {REPORT_NAME} files were found in:\n{in_dir}\n\n"
                     f"Put the district Highway Log PDFs (e.g. "
                     f"D01_Highway_Log_TSN.pdf) there, then run again."),
        )

    # Confirm overwrite *before* spending time parsing PDFs.
    if out.exists() and not confirm(out):
        return ConsolidateResult(status="cancelled",
                                 message="Cancelled. Existing file kept.")

    events.on_log("=" * 60)
    events.on_log(f"TSN Highway Log Conversion - {len(pdfs)} district PDF(s)")
    events.on_log("=" * 60)
    events.on_log("")

    # The combined workbook reflects exactly THIS run's PDFs: clear previously
    # converted files so districts removed from the input folder don't linger.
    CONVERTED_DIR.mkdir(parents=True, exist_ok=True)
    stale = list(CONVERTED_DIR.glob("tsn_highway_log_*.xlsx"))
    for p in stale:
        try:
            p.unlink()
        except OSError:
            return ConsolidateResult(
                status="error",
                message=(f"Could not replace {p.name}.\n\n"
                         "The file is probably open in Excel. Close it and try again."),
            )
    if stale:
        events.on_log(f"Cleared {len(stale)} previously converted file(s).")

    converted = 0
    total_rows = 0
    failed = []
    written = set()                  # guard against duplicate district+route across PDFs
    for i, p in enumerate(pdfs, 1):
        if events.is_cancelled():
            return ConsolidateResult(status="cancelled", message="Cancelled by user.")
        prefix = f"[{i}/{len(pdfs)}] {p.name}"
        events.on_log(f"{prefix} parsing…")
        try:
            district, route_rows = parse_pdf(str(p), events, pdf_name=p.name)
        except Exception as e:
            events.on_log(f"{prefix} FAILED ({type(e).__name__}): {e}")
            failed.append(p.name)
            continue
        if route_rows is None:                       # cancelled mid-PDF
            return ConsolidateResult(status="cancelled", message="Cancelled by user.")
        if not route_rows:
            events.on_log(f"{prefix} no highway-log data found; skipping")
            failed.append(p.name)
            continue
        if district is None:                         # last resort: the filename
            m = DISTRICT_FROM_NAME.search(p.stem)
            district = m.group(1).zfill(2) if m else "00"
        for route, rows in route_rows.items():
            out_file = CONVERTED_DIR / f"tsn_highway_log_d{district}_route_{route}.xlsx"
            if out_file.name in written:
                events.on_log(f"  WARNING: district {district} route {route} already "
                              f"converted from an earlier PDF; {p.name} replaces it "
                              "(is the same district in the folder twice?)")
            written.add(out_file.name)
            try:
                _write_route_workbook(rows, out_file)
            except PermissionError:
                return ConsolidateResult(
                    status="error",
                    message=(f"Could not save {out_file.name}.\n\n"
                             "The file is probably open in Excel. Close it and try again."),
                )
            events.on_log(f"  district {district} route {route}: {len(rows)} rows "
                          f"-> {out_file.name}")
            converted += 1
            total_rows += len(rows)

    if converted == 0:
        return ConsolidateResult(
            status="error",
            message=(f"None of the PDFs in:\n{in_dir}\n\ncontained readable "
                     f"{REPORT_NAME} data. Are they the TSN California State "
                     "Highway Log PDFs?"),
        )

    events.on_log("")

    # Combine all converted per-route files with the shared XLSX core (header
    # lock-in, Route column from the filename, streaming write). Overwrite was
    # already confirmed above.
    result = consolidate_xlsx(
        input_dir=CONVERTED_DIR, out_path=out, sheet_name=SHEET_NAME,
        report_name=REPORT_NAME, title="TSN Highway Log Consolidation",
        events=events, confirm_overwrite=lambda _p: True,
    )
    if result.status == "ok":
        result.summary_lines = [
            f"District PDFs:  {len(pdfs) - len(failed)} converted"
            + (f", {len(failed)} failed {failed}" if failed else ""),
            f"Route files:    {converted} (in {CONVERTED_DIR})",
        ] + result.summary_lines
    return result


if __name__ == "__main__":
    from cli import run_consolidate_cli
    run_consolidate_cli(consolidate)
