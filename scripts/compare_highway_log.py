"""Compare the consolidated TSN Highway Log against the consolidated TSMIS
Highway Log and report where the vendor data does NOT represent TSMIS.

Reads two workbooks (by default from output/):
    tsn_highway_log_consolidated.xlsx          (from the TSN report type here)
    [tsmis_]highway_log_consolidated.xlsx      (from the Highway Log report type)
Writes one workbook:  output/highway_log_comparison.xlsx

The whole point is to surface REAL discrepancies without calling out
differences that are merely representational. The comparison is therefore
structured around how the two systems actually encode the same highway:

  * Postmiles reset at county lines, so rows are only unique within a county
    section. Sections are detected by county-odometer resets and matched
    between the files by postmile overlap -- the files order counties
    differently (TSN by district file, TSMIS geographically), so order is
    never assumed.
  * TSN subdivides segments more finely than TSMIS (bridges, DVMS volume
    stations). The vendor is therefore judged AT TSMIS BREAKPOINTS: for every
    TSMIS row, the TSN row covering that postmile must carry the same
    attributes. TSN-only breakpoints that introduce no attribute change are
    counted as segmentation, never as discrepancies.
  * Representational conventions verified against the data and suppressed:
      - A blank TSMIS cell makes no claim, so it cannot be misrepresented:
        TSN values where TSMIS is blank are reported on their own
        informational sheet (TSN prints epoch dates like 640101, explicit
        'none' codes and mirrored roadbeds where TSMIS leaves blanks),
        never as discrepancies. The reverse -- TSMIS states a value and TSN
        drops or changes it -- IS a finding.
      - TSN prints "Sig Chg. Date" only when it differs from "Date of Rec"
        (0 of 26k TSN rows have them equal; TSMIS has thousands equal).
      - Med Wid leading zeros ("0Z" vs "00Z").
      - Whitespace padding, empty-vs-None.
      - "(DVMS) ..." volume-station annotations TSN appends to descriptions,
        and the different multi-line description encodings of the two systems.
      - MI (segment length) is only compared when both files break at the
        same two postmiles; otherwise it differs by segmentation, not data.
      - County-odometer differences are collapsed into constant-offset runs
        (one finding per run, not one per row).
  * Rows TSMIS dates AFTER the newest date present anywhere in the TSN file
    changed after the TSN snapshot was taken; their differences are reported
    on a separate sheet, not as vendor discrepancies.

Console-free like the other report modules: progress via events.on_log,
overwrite via the callback, cancel honored between routes, ConsolidateResult
returned. The console UX lives in cli.run_consolidate_cli.
"""
import collections
import re
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

from events import ConsolidateResult, Events
from paths import OUTPUT_ROOT

# Both inputs are products of this app, read back from output/ by default.
INPUT_DIR = OUTPUT_ROOT
OUT_PATH = OUTPUT_ROOT / "highway_log_comparison.xlsx"
TSN_PATTERN = "tsn_highway_log_consolidated*.xlsx"
TSMIS_PATTERNS = ("tsmis_highway_log_consolidated*.xlsx", "highway_log_consolidated*.xlsx")

REPORT_NAME = "Highway Log Comparison"
SHEET = "Highway Log"

# File pattern the GUI uses to preview how many inputs a folder holds.
INPUT_GLOB = "*highway_log_consolidated*.xlsx"

# Attribute columns judged at TSMIS breakpoints. (Location/MI/Cnty Odom and
# Description are handled with their own rules; see _compare.)
ATTR_COLS = ["N/A", "City", "R/U", "SPD", "TER", "H/G", "A/C",
             "LB T", "LB Lns", "LB F", "LB OT", "LB TR", "LB T-W", "LB IN", "LB SH",
             "Med TCB", "Med Wid", "RB T", "RB Lns", "RB F", "RB IN", "RB SH",
             "RB T-W", "RB OT", "RB SH2", "Date of Rec", "Sig Chg. Date"]

PM_NUM = re.compile(r"\d{3}\.\d{3}")
DVMS = re.compile(r"[,/ ]*\(?DVMS\)?[ 0-9,]*")


# =============================================================================
# Loading / structuring
# =============================================================================

def _load(path):
    """Load a consolidated Highway Log workbook -> (header, rows). The second
    'RB SH' column is renamed 'RB SH2' so columns can be addressed by name."""
    wb = load_workbook(path, read_only=True)
    if SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"{Path(path).name} has no '{SHEET}' sheet")
    rows_iter = wb[SHEET].iter_rows(values_only=True)
    hdr = list(next(rows_iter))
    seen = set()
    for i, h in enumerate(hdr):
        while h in seen:
            h = f"{h}2"
        hdr[i] = h
        seen.add(h)
    n = len(hdr)
    data = [list(r)[:n] + [None] * (n - len(r)) for r in rows_iter]
    wb.close()
    return hdr, data


def _pm(loc):
    """Numeric postmile from a Location like 'R012.887' / '026.437E'."""
    m = PM_NUM.search(str(loc) if loc is not None else "")
    return float(m.group(0)) if m else None


LOC_PARTS = re.compile(r"^([A-Z]?)(\d{3}\.\d{3})([A-Z]?)$")


def _suffix(loc):
    """Trailing alignment/equation letter of a Location ('' if none)."""
    m = LOC_PARTS.fullmatch(str(loc) if loc is not None else "")
    return m.group(3) if m else ""


def _is_alignment_row(loc):
    """True for TSMIS's separate left/right alignment series ('041.840L',
    'R081.505L', '012.220R'): TSMIS gives a diverged roadbed its own postmile
    rows, while TSN always folds both roadbeds into one row's LB/RB columns.
    These rows have no row-level TSN counterpart by design."""
    return _suffix(loc) in ("L", "R")


def _split_sections(rows, i_odom):
    """Split one route's rows into county sections at odometer resets."""
    sections, prev = [], None
    for r in rows:
        try:
            f = float(r[i_odom])
        except (TypeError, ValueError):
            f = None
        if not sections or (f is not None and prev is not None and f < prev - 0.001):
            sections.append([])
        if f is not None:
            prev = f
        sections[-1].append(r)
    return sections


def _match_sections(t_secs, n_secs, i_loc):
    """Pair TSMIS sections with TSN sections by postmile-multiset overlap.
    Returns ([(t_idx, n_idx | None)], unmatched_tsn_idx). Order-free: the two
    files list counties in different orders."""
    def locs(sec):
        return collections.Counter(r[i_loc] for r in sec)
    lt, ln = [locs(s) for s in t_secs], [locs(s) for s in n_secs]
    pairs, used = [], set()
    for i in sorted(range(len(t_secs)), key=lambda i: -len(t_secs[i])):
        best, bj = 0, None
        for j in range(len(n_secs)):
            if j in used:
                continue
            ov = sum((lt[i] & ln[j]).values())
            if ov > best:
                best, bj = ov, j
        if bj is not None and best >= max(2, 0.2 * len(t_secs[i])):
            used.add(bj)
            pairs.append((i, bj))
        else:
            pairs.append((i, None))
    pairs.sort()
    return pairs, [j for j in range(len(n_secs)) if j not in used]


# =============================================================================
# Normalization (the false-positive guards)
# =============================================================================

def _norm(col, v):
    """Representation-neutral form of one cell value."""
    if v is None:
        return ""
    s = re.sub(r"\s+", " ", str(v)).strip()
    if col == "Med Wid":
        # TSMIS prints '0Z' where TSN prints '00Z' -- same width, same code.
        m = re.fullmatch(r"(\d+)([A-Z]*)", s)
        if m:
            s = str(int(m.group(1))) + m.group(2)
    return s


def _norm_desc(v, tsn_side):
    """Descriptions: both systems wrap long text differently (TSMIS pads parts
    with spaces and joins with ','; TSN prints separate lines / '/'), and TSN
    appends '(DVMS) nnn,nnn' volume-station notes. Compare letters+digits only."""
    if v is None:
        return ""
    s = str(v)
    if tsn_side:
        s = DVMS.sub(" ", s)
    return re.sub(r"[^A-Z0-9]", "", s.upper())


def _date_key(v):
    """Sortable (year, mmdd) from a 6-digit yymmdd; None if not a date."""
    s = str(v) if v is not None else ""
    if not re.fullmatch(r"\d{6}", s):
        return None
    yy = int(s[:2])
    return (2000 + yy if yy < 50 else 1900 + yy, s[2:])


# =============================================================================
# Comparison engine
# =============================================================================

def _medtcb_note(va, vb):
    """TSMIS sometimes shows a partial median code ('07' or '7Z') where TSN
    shows a full 3-character one ('B7Z'). Tag ranges where the shared part
    agrees so reviewers can separate format truncation from real differences."""
    if re.fullmatch(r"\d[A-Z]", va) and len(vb) == 3 and vb.endswith(va):
        return "TSMIS shows partial code; shared part agrees"
    if re.fullmatch(r"0\d", va) and len(vb) == 3 and vb[1] == va[1]:
        return "TSMIS shows partial code; shared digit agrees"
    return ""


def _merge_runs(cells):
    """Merge per-breakpoint cell findings into postmile ranges.

    cells: ordered [(bp_ordinal, pm_from, pm_to, col, raw_t, raw_n, norm_key,
    exact)] for ONE (route, section). Consecutive breakpoints with the same
    column and the same normalized value pair become one range."""
    by_col = collections.defaultdict(list)
    for cell in cells:
        by_col[cell[3]].append(cell)
    out = []
    for col, items in by_col.items():
        run = None
        for bp, pf, pt, _c, vt, vn, key, exact in items:
            if run and key == run["key"] and bp == run["bp"] + 1:
                run["pm_to"] = pt
                run["bp"] = bp
                run["n"] += 1
                run["n_exact"] += 1 if exact else 0
            else:
                if run:
                    out.append(run)
                run = {"key": key, "col": col, "pm_from": pf, "pm_to": pt,
                       "bp": bp, "n": 1, "n_exact": 1 if exact else 0,
                       "vt": vt, "vn": vn}
        if run:
            out.append(run)
    out.sort(key=lambda r: (r["pm_from"] if r["pm_from"] is not None else -1, r["col"]))
    return out


def _run_row(route, sec_label, run):
    note = ""
    if run["col"] == "Med TCB":
        note = _medtcb_note(run["key"][0], run["key"][1])
    return [route, sec_label, run["pm_from"], run["pm_to"], run["n"],
            f'{run["n_exact"]}/{run["n"]}', run["col"], run["vt"], run["vn"], note]


def _compare(hdr, tsmis, tsn, events):
    """Run the comparison. Returns a dict of result tables, or None on cancel."""
    I = {h: i for i, h in enumerate(hdr)}
    i_loc, i_odom, i_mi = I["Location"], I["Cnty Odom"], I["MI"]
    i_desc, i_rec, i_sig = I["Description"], I["Date of Rec"], I["Sig Chg. Date"]

    by_route_t = collections.defaultdict(list)
    by_route_n = collections.defaultdict(list)
    for r in tsmis:
        by_route_t[r[0]].append(r)
    for r in tsn:
        by_route_n[r[0]].append(r)

    # Newest date anywhere in the TSN file = the vendor snapshot horizon.
    tsn_cutoff = max((k for r in tsn for k in (_date_key(r[i_rec]), _date_key(r[i_sig])) if k),
                     default=(9999, ""))

    res = {
        "diffs": [],            # Route, Section, PM from/to, rows, exact, Column, TSMIS, TSN, Note
        "post_snapshot": [],    # same shape (TSMIS row dated after the TSN snapshot)
        "tsn_fills": [],        # same shape (TSN value where TSMIS is blank)
        "tsn_plus": [],         # same shape (TSN prints +/++ placeholders)
        "odometer": [],         # Route, Section, PM from, PM to, rows, TSN minus TSMIS
        "desc": [],             # Route, Section, Location, TSMIS, TSN
        "uncovered": [],        # Route, Section, Location (no TSN row at/before pm)
        "extra_rows": [],       # Route, Section, Location, Description (TSN adds content)
        "sections_missing": [], # Route, PM range, rows
        "sections_extra": [],
        "alignments": [],       # Route, side, rows, PM range (L/R-suffix series)
        "routes_tsmis_only": sorted(set(by_route_t) - set(by_route_n)),
        "routes_tsn_only": sorted(set(by_route_n) - set(by_route_t)),
        "counts": collections.Counter(),
        "col_counts": collections.Counter(),
        "tsn_cutoff": tsn_cutoff,
    }
    cnt = res["counts"]
    cnt["tsmis_rows"] = len(tsmis)
    cnt["tsn_rows"] = len(tsn)
    common_routes = sorted(set(by_route_t) & set(by_route_n))

    for ri, route in enumerate(common_routes, 1):
        if events.is_cancelled():
            return None
        if ri % 40 == 0:
            events.on_log(f"  …route {route} ({ri}/{len(common_routes)})")
        # TSMIS represents diverged left/right roadbeds as their own L/R-suffix
        # postmile series; TSN folds both roadbeds into one row. Those rows are
        # structurally non-comparable row-by-row: set them aside, explained.
        rows_t, rows_n = [], []
        align_pms = set()
        for src, dst, side in ((by_route_t[route], rows_t, "TSMIS"),
                               (by_route_n[route], rows_n, "TSN")):
            align = []
            for r in src:
                (align if _is_alignment_row(r[i_loc]) else dst).append(r)
            if align:
                res["alignments"].append(
                    [route, side, len(align),
                     f"{align[0][i_loc]} – {align[-1][i_loc]}"])
                cnt["alignment_rows_" + side.lower()] += len(align)
                if side == "TSMIS":
                    align_pms |= {_pm(r[i_loc]) for r in align}

        # ------------------------------------------------------------------
        # ROW PAIRING. The two files put county/realignment odometer resets in
        # different places, so section structure is NOT trusted for pairing.
        #  pass 1: exact (Location, Cnty Odom), unique on both sides -- the
        #          same physical point, certain.
        #  pass 2: unique Location among the leftovers -- same point, the
        #          odometer drifted (which is itself reported).
        # Duplicated keys (e.g. '000.000' at several county starts) fall back
        # to the section machinery below, which pairs them county-correctly.
        # ------------------------------------------------------------------
        def k_full(r):
            return (str(r[i_loc]), str(r[i_odom]))
        c_t = collections.Counter(k_full(r) for r in rows_t)
        c_n = collections.Counter(k_full(r) for r in rows_n)
        n_by_key = {}
        for j, r in enumerate(rows_n):
            n_by_key.setdefault(k_full(r), []).append(j)
        pair = [None] * len(rows_t)               # tsmis idx -> tsn idx
        n_paired = [False] * len(rows_n)
        for i, r in enumerate(rows_t):
            k = k_full(r)
            if c_t[k] == 1 and c_n.get(k) == 1:
                j = n_by_key[k][0]
                pair[i] = j
                n_paired[j] = True
        lo_t = collections.Counter(str(rows_t[i][i_loc]) for i in range(len(rows_t))
                                   if pair[i] is None)
        lo_n = collections.Counter(str(rows_n[j][i_loc]) for j in range(len(rows_n))
                                   if not n_paired[j])
        n_by_loc = {}
        for j, r in enumerate(rows_n):
            if not n_paired[j]:
                n_by_loc.setdefault(str(r[i_loc]), []).append(j)
        for i, r in enumerate(rows_t):
            if pair[i] is not None:
                continue
            loc = str(r[i_loc])
            if lo_t[loc] == 1 and lo_n.get(loc) == 1:
                j = n_by_loc[loc][0]
                pair[i] = j
                n_paired[j] = True

        # Section structure (for run labels, covering and leftover pairing).
        t_secs = _split_sections(rows_t, i_odom)
        n_secs = _split_sections(rows_n, i_odom)
        sec_pairs, extra_secs = _match_sections(t_secs, n_secs, i_loc)
        # flat index ranges per section
        t_starts, n_starts = [], []
        pos = 0
        for s in t_secs:
            t_starts.append(pos)
            pos += len(s)
        pos = 0
        for s in n_secs:
            n_starts.append(pos)
            pos += len(s)
        # per-row "next postmile within my own section" (for the MI rule)
        next_pm_t = [None] * len(rows_t)
        next_pm_n = [None] * len(rows_n)
        for starts, secs, arr in ((t_starts, t_secs, next_pm_t),
                                  (n_starts, n_secs, next_pm_n)):
            for s0, sec in zip(starts, secs):
                nxt = None
                for off in range(len(sec) - 1, -1, -1):
                    arr[s0 + off] = nxt
                    pm = _pm(sec[off][i_loc])
                    if pm is not None:
                        nxt = pm
        n_sec_of = [None] * len(rows_n)
        for sj, (s0, sec) in enumerate(zip(n_starts, n_secs)):
            for off in range(len(sec)):
                n_sec_of[s0 + off] = sj

        # leftover pairing within matched sections (county-start duplicates):
        # key (pm, suffix, occurrence-among-leftovers)
        nj_of_ti = {ti: nj for ti, nj in sec_pairs if nj is not None}
        for ti, nj in nj_of_ti.items():
            def keyed_left(sec, s0, taken):
                seen = collections.Counter()
                out = {}
                for off, r in enumerate(sec):
                    if taken(s0 + off):
                        continue
                    pk = (_pm(r[i_loc]), _suffix(r[i_loc]))
                    out[(*pk, seen[pk])] = s0 + off
                    seen[pk] += 1
                return out
            kt = keyed_left(t_secs[ti], t_starts[ti], lambda i: pair[i] is not None)
            kn = keyed_left(n_secs[nj], n_starts[nj], lambda j: n_paired[j])
            for k, i in kt.items():
                j = kn.get(k)
                if j is not None:
                    pair[i] = j
                    n_paired[j] = True

        # judgment walk: TSMIS sections in order
        for ti, tsec in enumerate(t_secs):
            s0 = t_starts[ti]
            nj = nj_of_ti.get(ti)
            sec_label = f"{tsec[0][i_loc]}–{tsec[-1][i_loc]}"
            nsec = n_secs[nj] if nj is not None else None

            def covering(pm):
                """TSN row of the matched section active at postmile pm."""
                if nsec is None or pm is None:
                    return None
                best = None
                for r in nsec:
                    p = _pm(r[i_loc])
                    if p is None:
                        continue
                    if p <= pm + 1e-9:
                        best = r
                    else:
                        break
                return best

            cells, cells_post, cells_fill, cells_plus = [], [], [], []
            run_odo = None

            for bp, tr in enumerate(tsec):
                i = s0 + bp
                pm = _pm(tr[i_loc])
                j = pair[i]
                exact = j is not None
                nr = rows_n[j] if exact else covering(pm)
                cnt["tsmis_breakpoints"] += 1
                if nr is None:
                    res["uncovered"].append([route, sec_label, tr[i_loc]])
                    cnt["uncovered"] += 1
                    continue
                cnt["matched_breakpoints" if exact else "covered_breakpoints"] += 1

                post = any(d and d > tsn_cutoff
                           for d in (_date_key(tr[i_rec]), _date_key(tr[i_sig])))
                pm_next = next_pm_t[i] if next_pm_t[i] is not None else pm

                target = cells_post if post else cells
                for col in ATTR_COLS:
                    va, vb = _norm(col, tr[I[col]]), _norm(col, nr[I[col]])
                    if col == "Sig Chg. Date" and vb == "" and va != "" \
                            and _norm(col, tr[i_rec]) == va:
                        continue                     # TSN omits sig == rec by design
                    if va == vb:
                        continue
                    if va == "":
                        # blank TSMIS cell: nothing to misrepresent. TSN adding
                        # a value is informational, never a discrepancy.
                        cells_fill.append((bp, pm, pm_next, col,
                                           tr[I[col]], nr[I[col]], (va, vb), exact))
                        cnt["tsn_fill_cells"] += 1
                        continue
                    if vb and set(vb) == {"+"}:
                        # TSN prints +/++/+++ across a roadbed block: a report
                        # placeholder ("value not shown"), not a stated value.
                        cells_plus.append((bp, pm, pm_next, col,
                                           tr[I[col]], nr[I[col]], (va, vb), exact))
                        cnt["tsn_placeholder_cells"] += 1
                        continue
                    target.append((bp, pm, pm_next, col,
                                   tr[I[col]], nr[I[col]], (va, vb), exact))
                    if post:
                        cnt["post_snapshot_cells"] += 1
                    else:
                        cnt["diff_cells"] += 1
                        res["col_counts"][col] += 1

                if exact:
                    # Location string (realignment prefix / equation suffix)
                    if str(tr[i_loc]) != str(nr[i_loc]):
                        target.append((bp, pm, pm, "Location", tr[i_loc], nr[i_loc],
                                       (str(tr[i_loc]), str(nr[i_loc])), True))
                        cnt["post_snapshot_cells" if post else "diff_cells"] += 1
                        if not post:
                            res["col_counts"]["Location"] += 1
                    # MI only when the two files describe the IDENTICAL span:
                    # same start and both next breakpoints at the same postmile
                    # (a TSN split makes its MI legitimately shorter).
                    tn = next_pm_n[j]
                    if next_pm_t[i] is not None and tn is not None \
                            and abs(tn - next_pm_t[i]) < 1e-9:
                        va, vb = _norm("MI", tr[i_mi]), _norm("MI", nr[i_mi])
                        if va and vb and va != vb:
                            target.append((bp, pm, pm_next, "MI",
                                           tr[i_mi], nr[i_mi], (va, vb), True))
                            cnt["post_snapshot_cells" if post else "diff_cells"] += 1
                            if not post:
                                res["col_counts"]["MI"] += 1
                    # Description, representation-neutral
                    da, db = _norm_desc(tr[i_desc], False), _norm_desc(nr[i_desc], True)
                    if da != db:
                        res["desc"].append([route, sec_label, tr[i_loc],
                                            tr[i_desc], nr[i_desc]])
                        cnt["desc_diffs"] += 1
                    # Odometer offset runs
                    try:
                        delta = round(float(nr[i_odom]) - float(tr[i_odom]), 3)
                    except (TypeError, ValueError):
                        delta = None
                    if run_odo and delta == run_odo[2]:
                        run_odo[1] = pm
                        run_odo[3] += 1
                    else:
                        if run_odo and run_odo[2] not in (0.0, None):
                            res["odometer"].append([route, sec_label, run_odo[0],
                                                    run_odo[1], run_odo[3], run_odo[2]])
                            cnt["odometer_runs"] += 1
                        run_odo = [pm, pm, delta, 1]

            if run_odo and run_odo[2] not in (0.0, None):
                res["odometer"].append([route, sec_label, run_odo[0], run_odo[1],
                                        run_odo[3], run_odo[2]])
                cnt["odometer_runs"] += 1

            for run in _merge_runs(cells):
                res["diffs"].append(_run_row(route, sec_label, run))
            for run in _merge_runs(cells_post):
                res["post_snapshot"].append(_run_row(route, sec_label, run))
            for run in _merge_runs(cells_fill):
                res["tsn_fills"].append(_run_row(route, sec_label, run))
            for run in _merge_runs(cells_plus):
                res["tsn_plus"].append(_run_row(route, sec_label, run))

        # --- unpaired TSN rows: segmentation vs extra content ---------------
        for sj, (s0, nsec) in enumerate(zip(n_starts, n_secs)):
            for off, nr in enumerate(nsec):
                if n_paired[s0 + off]:
                    continue
                prev = nsec[off - 1] if off > 0 else None
                if prev is not None and all(
                        _norm(col, nr[I[col]]) == _norm(col, prev[I[col]])
                        for col in ATTR_COLS if col not in ("Date of Rec", "Sig Chg. Date")):
                    cnt["tsn_segmentation_rows"] += 1   # a pure split; no new claim
                else:
                    res["extra_rows"].append(
                        [route, f"{nsec[0][i_loc]}–{nsec[-1][i_loc]}",
                         nr[i_loc], nr[i_desc] or ""])
                    cnt["tsn_extra_rows"] += 1

        # --- section coverage: judged on PAIRING, not on section matching ---
        for ti, tsec in enumerate(t_secs):
            s0 = t_starts[ti]
            unpaired = sum(1 for off in range(len(tsec)) if pair[s0 + off] is None)
            if len(tsec) >= 2 and unpaired >= 0.7 * len(tsec) and nj_of_ti.get(ti) is None:
                res["sections_missing"].append(
                    [route, f"{tsec[0][i_loc]} – {tsec[-1][i_loc]}", len(tsec)])
                cnt["tsmis_unmatched_section_rows"] += len(tsec)
        for sj, (s0, nsec) in enumerate(zip(n_starts, n_secs)):
            unpaired = sum(1 for off in range(len(nsec)) if not n_paired[s0 + off])
            if len(nsec) >= 2 and unpaired >= 0.7 * len(nsec):
                pms = {_pm(r[i_loc]) for r in nsec}
                if pms and len(pms & align_pms) >= 0.3 * len(pms):
                    res["sections_extra"].append(
                        [route, f"{nsec[0][i_loc]} – {nsec[-1][i_loc]}", len(nsec),
                         "TSN alignment series (pairs with TSMIS L/R rows)"])
                    cnt["tsn_alignment_section_rows"] += len(nsec)
                else:
                    res["sections_extra"].append(
                        [route, f"{nsec[0][i_loc]} – {nsec[-1][i_loc]}", len(nsec), ""])
                    cnt["tsn_extra_section_rows"] += len(nsec)
    return res


# =============================================================================
# Report workbook
# =============================================================================

def _write_report(res, out_path, tsmis_name, tsn_name):
    header_fill = PatternFill("solid", start_color="305496")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    title_font = Font(name="Arial", bold=True, size=12)
    note_font = Font(name="Arial", size=9, color="595959")

    wb = Workbook()
    cnt = res["counts"]

    def sheet(name, headers, rows, widths):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for c in ws[1]:
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for r in rows:
            ws.append(r)
        return ws

    # --- Summary -------------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 64
    ws.column_dimensions["B"].width = 14
    y, mx = res["tsn_cutoff"], None
    cutoff = f"{y[0]}-{y[1][:2]}-{y[1][2:]}" if y and y[0] != 9999 else "n/a"

    def line(text, value=None, font=None):
        ws.append([text, value])
        if font:
            ws.cell(row=ws.max_row, column=1).font = font

    line("TSN vs TSMIS Highway Log comparison", None, title_font)
    line(f"TSMIS file: {tsmis_name}", None, note_font)
    line(f"TSN file:   {tsn_name}", None, note_font)
    line(f"Newest date in the TSN data (snapshot horizon): {cutoff}", None, note_font)
    line("")
    line("SCOPE", None, title_font)
    line("TSMIS rows", cnt["tsmis_rows"])
    line("TSN rows", cnt["tsn_rows"])
    line("Routes only in TSMIS (not judged)", len(res["routes_tsmis_only"]))
    line("Routes only in TSN (not judged)", len(res["routes_tsn_only"]))
    line("TSMIS breakpoints judged", cnt["tsmis_breakpoints"])
    line("  with an identical TSN breakpoint", cnt["matched_breakpoints"])
    line("  covered by an enclosing TSN segment", cnt["covered_breakpoints"])
    line("  with no TSN coverage", cnt["uncovered"])
    line("")
    line("FINDINGS (TSMIS states a value; TSN differs)", None, title_font)
    line("Attribute difference ranges (see 'Attribute differences')", len(res["diffs"]))
    line("  differing cells behind those ranges", cnt["diff_cells"])
    line("Odometer offset runs (see 'Odometer offsets')", cnt["odometer_runs"])
    n_align_secs = sum(1 for r in res["sections_extra"] if r[3])
    line("County sections missing in TSN", len(res["sections_missing"]))
    line("County sections only in TSN (excl. alignment series)",
         len(res["sections_extra"]) - n_align_secs)
    line("TSN rows adding content not in TSMIS (see 'Extra in TSN')", cnt["tsn_extra_rows"])
    line("TSMIS breakpoints with no TSN coverage", cnt["uncovered"])
    line("")
    line("INFORMATIONAL (not counted as discrepancies)", None, title_font)
    line("TSN values where TSMIS is blank (see its sheet)", cnt["tsn_fill_cells"])
    line("TSN '+' placeholders over TSMIS values (see its sheet)",
         cnt["tsn_placeholder_cells"])
    line("Description wording differences", cnt["desc_diffs"])
    line("Changed in TSMIS after the TSN snapshot (see 'Post-snapshot')",
         len(res["post_snapshot"]))
    line("TSN-only segment splits with identical attributes (suppressed)",
         cnt["tsn_segmentation_rows"])
    line("TSMIS separate left/right alignment rows (not row-comparable)",
         cnt["alignment_rows_tsmis"])
    line("TSN alignment-series sections (TSN's copy of those roadbeds)",
         n_align_secs)
    line("")
    line("RULES APPLIED (to avoid false discrepancies)", None, title_font)
    for t in (
        "Counties are matched by postmile overlap; file order is ignored.",
        "The vendor is judged at TSMIS breakpoints; TSN's finer segmentation "
        "(bridges, DVMS volume stations) is never counted as a difference.",
        "A blank TSMIS cell makes no claim: TSN values where TSMIS is blank "
        "are informational, never discrepancies. TSMIS values that TSN drops "
        "or changes ARE findings.",
        "TSN prints Sig Chg. Date only when it differs from Date of Rec; such "
        "rows are treated as equal.",
        "MI is compared only where both files describe the identical span "
        "(same start and end, no TSN split in between).",
        "TSN rows printed as '+/++' placeholders (value not shown by the TSN "
        "report) are informational, not value contradictions.",
        "Med Wid leading zeros, whitespace and empty-vs-blank are ignored.",
        "Descriptions are compared on letters and digits only; TSN '(DVMS)' "
        "annotations are removed first.",
        "TSMIS rows dated after the TSN snapshot are reported separately.",
        "TSMIS describes diverged left/right roadbeds as separate 'L'/'R'-"
        "suffixed postmile series; TSN folds both roadbeds into one row, so "
        "those rows are listed (see 'Separate alignments'), not compared.",
        "'Common bps' shows how many breakpoints in a range exist identically "
        "in both files (the rest are TSMIS points covered by a TSN segment).",
    ):
        line("• " + t, None, note_font)

    pc = ws.max_row + 2
    ws.cell(row=pc, column=1, value="DIFFERING CELLS BY COLUMN").font = title_font
    for col, n in res["col_counts"].most_common():
        ws.append([col, n])

    # --- detail sheets ---------------------------------------------------------
    FIND_HDR = ["Route", "County section (PM range)", "PM from", "PM to",
                "TSMIS rows", "Common bps", "Column", "TSMIS value", "TSN value", "Note"]
    FIND_W = [8, 24, 10, 10, 10, 11, 12, 20, 20, 40]
    sheet("Attribute differences", FIND_HDR, res["diffs"], FIND_W)
    sheet("Odometer offsets",
          ["Route", "County section (PM range)", "PM from", "PM to",
           "Breakpoints", "TSN odometer minus TSMIS (mi)"],
          res["odometer"], [8, 24, 10, 10, 12, 26])
    sheet("Extra in TSN",
          ["Route", "County section (PM range)", "Location", "Description"],
          res["extra_rows"], [8, 24, 12, 60])
    sheet("No TSN coverage",
          ["Route", "County section (PM range)", "Location"],
          res["uncovered"], [8, 24, 12])
    sheet("Sections coverage",
          ["Issue", "Route", "PM range", "Rows", "Note"],
          [["Missing in TSN", *r, ""] for r in res["sections_missing"]] +
          [["Only in TSN", *r] for r in res["sections_extra"]],
          [16, 8, 28, 8, 44])
    sheet("Route coverage",
          ["Issue", "Route"],
          [["Only in TSMIS", r] for r in res["routes_tsmis_only"]] +
          [["Only in TSN", r] for r in res["routes_tsn_only"]],
          [16, 10])
    sheet("Separate alignments (L-R)",
          ["Route", "File", "Rows", "PM range"],
          res["alignments"], [8, 10, 8, 28])
    sheet("TSN value where TSMIS blank", FIND_HDR, res["tsn_fills"], FIND_W)
    sheet("TSN placeholder (+)", FIND_HDR, res["tsn_plus"], FIND_W)
    sheet("Description differences",
          ["Route", "County section (PM range)", "Location", "TSMIS", "TSN"],
          res["desc"], [8, 24, 12, 55, 55])
    sheet("Post-snapshot", FIND_HDR, res["post_snapshot"], FIND_W)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# =============================================================================
# Entry point
# =============================================================================

def _find_inputs(folder):
    """Locate the TSN and TSMIS consolidated workbooks in `folder`."""
    tsn = sorted(folder.glob(TSN_PATTERN))
    tsmis = []
    for pat in TSMIS_PATTERNS:
        tsmis += [p for p in folder.glob(pat) if not p.name.startswith("tsn_")]
    tsmis = sorted(set(tsmis))
    return (tsn[-1] if tsn else None), (tsmis[-1] if tsmis else None)


def consolidate(events=None, confirm_overwrite=None, input_dir=None, out_path=None):
    """Compare the two consolidated Highway Log workbooks and write the
    comparison report. Fits the standard report contract so the GUI and the
    console menu can run it like any other report type."""
    in_dir = Path(input_dir) if input_dir else INPUT_DIR
    out = Path(out_path) if out_path else OUT_PATH
    events = events or Events()
    if not _DEPS_OK:
        return ConsolidateResult(
            status="error",
            message="Required components are missing (openpyxl).",
        )
    confirm = confirm_overwrite or (lambda _p: True)

    tsn_path, tsmis_path = _find_inputs(in_dir) if in_dir.exists() else (None, None)
    if tsn_path is None or tsmis_path is None:
        return ConsolidateResult(
            status="error",
            message=(f"Could not find both consolidated Highway Log workbooks in:\n{in_dir}\n\n"
                     "Needed: tsn_highway_log_consolidated.xlsx (run the TSN Highway Log "
                     "report) and [tsmis_]highway_log_consolidated.xlsx (run the Highway "
                     "Log report), or point the folder picker at where they are."),
        )

    if out.exists() and not confirm(out):
        return ConsolidateResult(status="cancelled", message="Cancelled. Existing file kept.")

    events.on_log("=" * 60)
    events.on_log("Highway Log Comparison — TSN vs TSMIS")
    events.on_log("=" * 60)
    events.on_log(f"TSMIS: {tsmis_path.name}")
    events.on_log(f"TSN:   {tsn_path.name}")
    events.on_log("")

    try:
        hdr_t, tsmis = _load(tsmis_path)
        hdr_n, tsn = _load(tsn_path)
    except Exception as e:
        return ConsolidateResult(status="error",
                                 message=f"Could not read the input workbooks: {e}")
    if hdr_t != hdr_n:
        return ConsolidateResult(
            status="error",
            message=("The two workbooks have different column layouts -- re-create "
                     "both with this app, then compare."))

    events.on_log(f"Comparing {len(tsmis)} TSMIS rows against {len(tsn)} TSN rows…")
    res = _compare(hdr_t, tsmis, tsn, events)
    if res is None:
        return ConsolidateResult(status="cancelled", message="Cancelled by user.")

    events.on_log("")
    events.on_log("Writing comparison workbook…")
    try:
        _write_report(res, out, tsmis_path.name, tsn_path.name)
    except PermissionError:
        return ConsolidateResult(
            status="error",
            message=(f"Could not save {out.name}.\n\n"
                     "The file is probably open in Excel. Close it and try again."),
        )

    cnt = res["counts"]
    return ConsolidateResult(
        status="ok",
        output_path=str(out),
        summary_lines=[
            f"TSMIS breakpoints judged: {cnt['tsmis_breakpoints']}",
            f"Attribute difference ranges: {len(res['diffs'])} "
            f"({cnt['diff_cells']} cells)",
            f"Odometer offset runs:     {cnt['odometer_runs']}",
            f"Sections missing/extra:   {len(res['sections_missing'])} / "
            f"{len(res['sections_extra'])}",
            f"Extra TSN content rows:   {cnt['tsn_extra_rows']} "
            f"(plus {cnt['tsn_segmentation_rows']} pure splits, suppressed)",
            f"Changed after TSN snapshot: {len(res['post_snapshot'])} ranges",
            f"Output file: {out}",
        ],
    )


if __name__ == "__main__":
    from cli import run_consolidate_cli
    run_consolidate_cli(consolidate)
