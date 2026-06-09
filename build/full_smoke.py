"""Comprehensive runtime self-test for the bundled libraries.

Exercises EVERY real code path the app depends on -- pdfplumber text/word
extraction (what the Ramp Summary consolidator uses), an openpyxl write/read
round-trip, a REAL consolidator run over synthetic per-route XLSX inputs --
then reports which *optional* libraries actually got imported. Used two ways:
  1. Against the build venv, to prove PIL/pypdfium2 are never loaded (so they
     can be excluded from the bundle).
  2. Frozen (built by build.ps1 -SelfTest), as the gate that proves a pruned
     bundle still runs everything.

No browser and no network: the test PDF is written byte-by-byte right here.

Exit 0 = all good. Nonzero/raise = something the app needs is broken.
"""
import sys
import tempfile
from pathlib import Path

# Make the app modules importable (frozen builds bundle them; dev/venv runs
# need the repo on sys.path).
if not getattr(sys, "frozen", False):
    _repo = Path(__file__).resolve().parents[1]
    sys.path.insert(0, str(_repo / "scripts"))      # events, gui_app, ...
    sys.path.insert(0, str(_repo))                  # version.py at repo root

import openpyxl                                   # noqa: E402
import pdfplumber                                 # noqa: E402


def write_min_pdf(path, text="Route 005 Ramp Summary 1234"):
    """Write a minimal one-page PDF (Helvetica text object) by hand, so the
    pdfplumber path can be exercised without any browser to print one."""
    objs = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
        b"/Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>",
        None,  # content stream, built below
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    ]
    stream = f"BT /F1 12 Tf 72 720 Td ({text}) Tj ET".encode("ascii")
    objs[3] = (b"<< /Length " + str(len(stream)).encode() + b" >>\nstream\n"
               + stream + b"\nendstream")

    out = bytearray(b"%PDF-1.4\n")
    offsets = []
    for i, body in enumerate(objs, 1):
        offsets.append(len(out))
        out += f"{i} 0 obj\n".encode() + body + b"\nendobj\n"
    xref_pos = len(out)
    out += f"xref\n0 {len(objs) + 1}\n".encode()
    out += b"0000000000 65535 f \n"
    for off in offsets:
        out += f"{off:010d} 00000 n \n".encode()
    out += (f"trailer\n<< /Size {len(objs) + 1} /Root 1 0 R >>\n"
            f"startxref\n{xref_pos}\n%%EOF\n").encode()
    Path(path).write_bytes(bytes(out))


def main() -> int:
    tmp = Path(tempfile.mkdtemp())
    print("=" * 60)
    print("TSMIS Consolidator -- full bundle self-test")
    print("=" * 60)
    print(f"frozen={getattr(sys, 'frozen', False)}  "
          f"openpyxl={openpyxl.__version__}  pdfplumber={pdfplumber.__version__}")

    # 1. pdfplumber: the calls consolidate_ramp_summary makes, over a PDF
    #    written by hand (no browser in this app to print one).
    pdf_path = tmp / "page.pdf"
    write_min_pdf(pdf_path)
    with pdfplumber.open(str(pdf_path)) as pdf:
        text = pdf.pages[0].extract_text() or ""
        words = pdf.pages[0].extract_words()
    assert "Route 005" in text, f"extract_text failed: {text!r}"
    assert any(w.get("text") == "1234" for w in words), "extract_words failed"
    print(f"pdfplumber: text={len(text)} chars, words={len(words)}")

    # 2. openpyxl: write + read round-trip (consolidator output path).
    xlsx = tmp / "wb.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Route", "Ramp", "Count"]); ws.append(["005", "NB On", 1234])
    wb.save(str(xlsx))
    wb2 = openpyxl.load_workbook(str(xlsx))
    assert wb2.active["C2"].value == 1234, "openpyxl round-trip failed"
    print("openpyxl: write/read round-trip ok")

    # 3. A REAL consolidator run over synthetic per-route inputs, end to end
    #    (header lock-in, route extraction, write_only streaming, styling).
    from events import Events
    import consolidate_highway_log as chl

    in_dir = tmp / "highway_log"; in_dir.mkdir()
    for route in ("005", "099"):
        wbr = openpyxl.Workbook(); wsr = wbr.active
        wsr.title = chl.SHEET_NAME
        wsr.append(["County", "Postmile"]); wsr.append(["LA", 1.5])
        wbr.save(str(in_dir / f"highway_log_route_{route}.xlsx"))
    out_path = tmp / "consolidated.xlsx"
    result = chl.consolidate(events=Events(), input_dir=in_dir, out_path=out_path)
    assert result.status == "ok", f"consolidate failed: {result.message}"
    wbc = openpyxl.load_workbook(str(out_path))
    rows = list(wbc[chl.SHEET_NAME].iter_rows(values_only=True))
    assert rows[0][:2] == ("Route", "County") and len(rows) == 3, f"bad output: {rows}"
    print("consolidator: 2 synthetic routes combined ok")

    # 4. Report optional libraries that should NOT be needed.
    opt = {m: (m in sys.modules) for m in ("PIL", "pypdfium2", "pypdfium2_raw")}
    print(f"optional libs loaded: {opt}")
    print(f"cryptography loaded (required by pdfminer): {'cryptography' in sys.modules}")

    # 5. GUI + app modules: construct the real window (withdrawn) and tear it
    #    down, so the self-test also catches a prune/exclude that broke an
    #    import the GUI needs. tkinter missing is only tolerable in a dev run
    #    (e.g. a slim Linux container); a frozen bundle without it is broken.
    try:
        import tkinter as tk
    except ImportError:
        if getattr(sys, "frozen", False):
            raise
        tk = None
        print("gui: skipped, tkinter not available in this dev environment")
    if tk is not None:
        try:
            import gui_app

            class _NoCheck:             # don't spawn the background check thread here
                def __init__(self, q): pass
                def start(self): pass
            gui_app.CheckWorker = _NoCheck

            app = gui_app.App()
            app.withdraw(); app.update_idletasks(); app.destroy()
            print("gui: App window constructed + torn down ok")
        except tk.TclError as e:
            print(f"gui: skipped, no display ({e})")

    import shutil
    shutil.rmtree(tmp, ignore_errors=True)
    print("\nSMOKE OK -- every app-required code path works.")
    # Signal to the caller (venv run) whether the excludable libs stayed out.
    if any(opt.values()):
        print(f"NOTE: optional libs were imported: "
              f"{[k for k, v in opt.items() if v]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
